"""XLSX generator.

Loads an uploaded .xlsx template and renders every string cell that contains a
Jinja2 marker (``{{`` or ``{%``) against the supplied content. Cells without a
marker are left untouched, so existing formatting, formulas and layout survive.

To repeat rows, drive a ``{% for %}`` loop from a list variable inside a cell —
do not paste the same placeholder into many cells (one variable renders to one
value everywhere). See README "Repeating content".
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from ._jinja import make_env, resolve_strict

_DEFAULT_STRICT = True


def _has_marker(value: str) -> bool:
    return "{{" in value or "{%" in value


def generate(template_bytes: bytes, content: dict, strict=None) -> bytes:
    env = make_env(resolve_strict(strict, _DEFAULT_STRICT))
    wb = load_workbook(BytesIO(template_bytes))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and _has_marker(cell.value):
                    cell.value = env.from_string(cell.value).render(**content)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
